"""
Tests for F018 — CSV/Excel import & export.

Coverage (25 tests):
 1.  parse_csv — basic
 2.  parse_csv — BOM-prefixed file
 3.  parse_csv — empty file returns empty list
 4.  parse_excel — basic
 5.  parse_excel — only-header workbook returns empty list
 6.  import_to_dimension — happy path (service layer)
 7.  import_to_dimension — skip duplicate name
 8.  import_to_dimension — parent column resolution
 9.  import_to_module — happy path (service layer)
10.  import_to_module — skips rows with no data
11.  export_module_to_csv — content check (service layer)
12.  export_module_to_excel — produces a valid workbook
13.  build_import_preview — column names and sample rows
14.  POST /dimensions/{id}/import — auth required
15.  POST /dimensions/{id}/import — CSV happy path
16.  POST /modules/{id}/import — auth required
17.  POST /modules/{id}/import — CSV happy path
18.  GET  /modules/{id}/export — auth required
19.  GET  /modules/{id}/export?format=csv — content check
20.  GET  /modules/{id}/export?format=xlsx — parseable output
21.  POST /modules/{id}/import/preview — auth required
22.  POST /modules/{id}/import/preview — happy path
23.  POST /dimensions/{id}/import — bad column name → 422
24.  POST /modules/{id}/import — no matching line-item columns → 422
25.  GET  /modules/{id}/export — unknown module → 404
"""
import csv
import io
import uuid

import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.import_export import (
    build_import_preview,
    export_module_to_csv,
    export_module_to_excel,
    import_to_dimension,
    import_to_module,
    parse_csv,
    parse_excel,
)


# ── Generic helpers ────────────────────────────────────────────────────────────

def _make_csv(rows: list, fieldnames: list) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames)
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _make_excel(rows: list, headers: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


async def _register_login(client: AsyncClient) -> str:
    email = f"u_{uuid.uuid4().hex[:8]}@test.com"
    pw = "Pass123!"
    await client.post("/auth/register", json={"email": email, "full_name": "T", "password": pw})
    r = await client.post("/auth/login", json={"email": email, "password": pw})
    assert r.status_code == 200
    return r.json()["access_token"]


async def _create_workspace(client: AsyncClient, token: str) -> str:
    r = await client.post("/workspaces/", json={"name": "WS"}, headers=_auth(token))
    assert r.status_code == 201
    return r.json()["id"]


async def _create_model(client: AsyncClient, token: str, ws_id: str) -> str:
    r = await client.post("/models", json={"name": "M", "workspace_id": ws_id}, headers=_auth(token))
    assert r.status_code == 201
    return r.json()["id"]


async def _create_dimension(client: AsyncClient, token: str, model_id: str) -> str:
    r = await client.post(
        f"/models/{model_id}/dimensions",
        json={"name": "Products", "dimension_type": "custom"},
        headers=_auth(token),
    )
    assert r.status_code == 201
    return r.json()["id"]


async def _create_module(client: AsyncClient, token: str, model_id: str) -> str:
    r = await client.post(
        f"/models/{model_id}/modules",
        json={"name": "Revenue"},
        headers=_auth(token),
    )
    assert r.status_code == 201
    return r.json()["id"]


async def _create_line_item(client: AsyncClient, token: str, module_id: str, name: str) -> str:
    r = await client.post(
        f"/modules/{module_id}/line-items",
        json={"name": name, "format": "number", "sort_order": 0},
        headers=_auth(token),
    )
    assert r.status_code == 201
    return r.json()["id"]


# ── DB session helper for service-layer tests ──────────────────────────────────

@pytest.fixture
async def db(client: AsyncClient):
    """Yield a live AsyncSession from the app's test DB override."""
    from app.core.database import get_db
    from app.main import app

    override = app.dependency_overrides.get(get_db)
    assert override is not None
    async for session in override():
        yield session
        return


# ── 1. parse_csv basic ────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_parse_csv_basic():
    content = _make_csv(
        [{"name": "Alpha", "code": "A"}, {"name": "Beta", "code": "B"}],
        ["name", "code"],
    )
    rows = parse_csv(content)
    assert len(rows) == 2
    assert rows[0]["name"] == "Alpha"
    assert rows[1]["code"] == "B"


# ── 2. parse_csv BOM ──────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_parse_csv_with_bom():
    raw = "name,code\r\nAlpha,A\r\n"
    bom_content = b"\xef\xbb\xbf" + raw.encode("utf-8")
    rows = parse_csv(bom_content)
    assert len(rows) == 1
    assert rows[0]["name"] == "Alpha"


# ── 3. parse_csv empty ────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_parse_csv_empty():
    rows = parse_csv(b"name,code\r\n")
    assert rows == []


# ── 4. parse_excel basic ──────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_parse_excel_basic():
    content = _make_excel(
        [{"product": "Alpha", "value": 100}, {"product": "Beta", "value": 200}],
        ["product", "value"],
    )
    rows = parse_excel(content)
    assert len(rows) == 2
    assert rows[0]["product"] == "Alpha"
    assert rows[1]["value"] == 200


# ── 5. parse_excel only-header ────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_parse_excel_only_header():
    content = _make_excel([], ["col1", "col2"])
    rows = parse_excel(content)
    assert rows == []


# ── 6. import_to_dimension happy path ────────────────────────────────────────

@pytest.mark.asyncio
async def test_import_to_dimension_happy_path(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    dim_id_str = await _create_dimension(client, token, model_id)
    dim_id = uuid.UUID(dim_id_str)

    rows = [{"name": "Widget A"}, {"name": "Widget B"}]
    result = await import_to_dimension(db, dim_id, rows, name_column="name")
    assert result.rows_imported == 2
    assert result.rows_skipped == 0


# ── 7. import_to_dimension skip duplicate ────────────────────────────────────

@pytest.mark.asyncio
async def test_import_to_dimension_skip_duplicate(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    dim_id_str = await _create_dimension(client, token, model_id)
    dim_id = uuid.UUID(dim_id_str)

    # First import
    rows = [{"name": "Widget A"}]
    await import_to_dimension(db, dim_id, rows, name_column="name")

    # Second import with same name
    result = await import_to_dimension(db, dim_id, rows, name_column="name")
    assert result.rows_skipped == 1
    assert any("already exists" in e for e in result.errors)


# ── 8. import_to_dimension with parent column ─────────────────────────────────

@pytest.mark.asyncio
async def test_import_to_dimension_with_parent(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    dim_id_str = await _create_dimension(client, token, model_id)
    dim_id = uuid.UUID(dim_id_str)

    # Import parent first
    await import_to_dimension(db, dim_id, [{"name": "Category"}], name_column="name")

    # Import child referencing parent
    child_rows = [{"name": "Sub-item", "parent": "Category"}]
    result = await import_to_dimension(
        db, dim_id, child_rows, name_column="name", parent_column="parent"
    )
    assert result.rows_imported == 1
    assert result.rows_skipped == 0


# ── 9. import_to_module happy path ────────────────────────────────────────────

@pytest.mark.asyncio
async def test_import_to_module_happy_path(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id_str = await _create_module(client, token, model_id)
    li1_id_str = await _create_line_item(client, token, module_id_str, "Sales")
    li2_id_str = await _create_line_item(client, token, module_id_str, "COGS")

    module_id = uuid.UUID(module_id_str)
    li1_id = uuid.UUID(li1_id_str)
    li2_id = uuid.UUID(li2_id_str)

    rows = [
        {"dim": "Jan", "Sales": "1000", "COGS": "400"},
        {"dim": "Feb", "Sales": "1200", "COGS": "500"},
    ]
    result = await import_to_module(db, module_id, rows, {"Sales": li1_id, "COGS": li2_id}, {})
    assert result.rows_imported == 2
    assert result.rows_skipped == 0


# ── 10. import_to_module skips rows with no data ─────────────────────────────

@pytest.mark.asyncio
async def test_import_to_module_skips_empty(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id_str = await _create_module(client, token, model_id)
    li1_id_str = await _create_line_item(client, token, module_id_str, "Sales")

    module_id = uuid.UUID(module_id_str)
    li1_id = uuid.UUID(li1_id_str)

    rows = [{"dim": "Jan", "Sales": None}]
    result = await import_to_module(db, module_id, rows, {"Sales": li1_id}, {})
    assert result.rows_skipped == 1
    assert result.rows_imported == 0


# ── 11. export_module_to_csv content check ────────────────────────────────────

@pytest.mark.asyncio
async def test_export_module_to_csv_content(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id_str = await _create_module(client, token, model_id)
    li1_id_str = await _create_line_item(client, token, module_id_str, "Sales")

    module_id = uuid.UUID(module_id_str)
    li1_id = uuid.UUID(li1_id_str)

    await import_to_module(db, module_id, [{"dim": "Jan", "Sales": "999"}], {"Sales": li1_id}, {})

    csv_bytes = await export_module_to_csv(db, module_id)
    assert isinstance(csv_bytes, bytes)
    text = csv_bytes.decode("utf-8")
    assert "line_item" in text
    assert "Sales" in text
    assert "999" in text


# ── 12. export_module_to_excel produces valid workbook ────────────────────────

@pytest.mark.asyncio
async def test_export_module_to_excel_valid(client: AsyncClient, db: AsyncSession):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id_str = await _create_module(client, token, model_id)
    li1_id_str = await _create_line_item(client, token, module_id_str, "Sales")

    module_id = uuid.UUID(module_id_str)
    li1_id = uuid.UUID(li1_id_str)

    await import_to_module(db, module_id, [{"dim": "Q1", "Sales": "5000"}], {"Sales": li1_id}, {})

    xlsx_bytes = await export_module_to_excel(db, module_id)
    assert isinstance(xlsx_bytes, bytes)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "line_item" in headers
    assert "value" in headers


# ── 13. build_import_preview ──────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_build_import_preview():
    rows = [
        {"Product": "Alpha", "Price": "10.0"},
        {"Product": "Beta", "Price": "20.0"},
        {"Product": "Gamma", "Price": "30.0"},
    ]
    preview = build_import_preview(rows, max_sample=2)
    assert preview["column_names"] == ["Product", "Price"]
    assert len(preview["sample_rows"]) == 2
    assert preview["suggested_mapping"] == {"Product": None, "Price": None}


# ── 14. POST /dimensions/{id}/import — auth required ─────────────────────────

@pytest.mark.asyncio
async def test_dimension_import_auth_required(client: AsyncClient):
    dim_id = uuid.uuid4()
    content = _make_csv([{"name": "X"}], ["name"])
    resp = await client.post(
        f"/dimensions/{dim_id}/import",
        files={"file": ("data.csv", content, "text/csv")},
    )
    assert resp.status_code == 401


# ── 15. POST /dimensions/{id}/import — CSV happy path ────────────────────────

@pytest.mark.asyncio
async def test_dimension_import_csv_happy_path(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    dim_id = await _create_dimension(client, token, model_id)

    content = _make_csv(
        [{"name": "Widget A"}, {"name": "Widget B"}],
        ["name"],
    )
    resp = await client.post(
        f"/dimensions/{dim_id}/import",
        files={"file": ("items.csv", content, "text/csv")},
        params={"name_column": "name"},
        headers=_auth(token),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["rows_imported"] == 2
    assert body["rows_skipped"] == 0


# ── 16. POST /modules/{id}/import — auth required ────────────────────────────

@pytest.mark.asyncio
async def test_module_import_auth_required(client: AsyncClient):
    module_id = uuid.uuid4()
    content = _make_csv([{"Sales": "100"}], ["Sales"])
    resp = await client.post(
        f"/modules/{module_id}/import",
        files={"file": ("data.csv", content, "text/csv")},
    )
    assert resp.status_code == 401


# ── 17. POST /modules/{id}/import — CSV happy path ───────────────────────────

@pytest.mark.asyncio
async def test_module_import_csv_happy_path(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id = await _create_module(client, token, model_id)
    await _create_line_item(client, token, module_id, "Sales")
    await _create_line_item(client, token, module_id, "COGS")

    content = _make_csv(
        [{"dim_key": "Jan", "Sales": "1000", "COGS": "400"}],
        ["dim_key", "Sales", "COGS"],
    )
    resp = await client.post(
        f"/modules/{module_id}/import",
        files={"file": ("data.csv", content, "text/csv")},
        headers=_auth(token),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["rows_imported"] == 1


# ── 18. GET /modules/{id}/export — auth required ──────────────────────────────

@pytest.mark.asyncio
async def test_module_export_auth_required(client: AsyncClient):
    module_id = uuid.uuid4()
    resp = await client.get(f"/modules/{module_id}/export")
    assert resp.status_code == 401


# ── 19. GET /modules/{id}/export?format=csv — content check ──────────────────

@pytest.mark.asyncio
async def test_module_export_csv_content(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id = await _create_module(client, token, model_id)
    await _create_line_item(client, token, module_id, "Sales")

    # Seed data
    content = _make_csv(
        [{"dim_key": "Jan", "Sales": "500"}],
        ["dim_key", "Sales"],
    )
    seed = await client.post(
        f"/modules/{module_id}/import",
        files={"file": ("data.csv", content, "text/csv")},
        headers=_auth(token),
    )
    assert seed.status_code == 200

    resp = await client.get(
        f"/modules/{module_id}/export",
        params={"format": "csv"},
        headers=_auth(token),
    )
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert "line_item" in resp.text


# ── 20. GET /modules/{id}/export?format=xlsx — parseable output ───────────────

@pytest.mark.asyncio
async def test_module_export_xlsx_parseable(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id = await _create_module(client, token, model_id)

    resp = await client.get(
        f"/modules/{module_id}/export",
        params={"format": "xlsx"},
        headers=_auth(token),
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.active is not None


# ── 21. POST /modules/{id}/import/preview — auth required ────────────────────

@pytest.mark.asyncio
async def test_preview_import_auth_required(client: AsyncClient):
    module_id = uuid.uuid4()
    content = _make_csv([{"Sales": "100"}], ["Sales"])
    resp = await client.post(
        f"/modules/{module_id}/import/preview",
        files={"file": ("data.csv", content, "text/csv")},
    )
    assert resp.status_code == 401


# ── 22. POST /modules/{id}/import/preview — happy path ───────────────────────

@pytest.mark.asyncio
async def test_preview_import_happy_path(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id = await _create_module(client, token, model_id)

    content = _make_csv(
        [{"dim": "Jan", "Sales": "100", "COGS": "40"}],
        ["dim", "Sales", "COGS"],
    )
    resp = await client.post(
        f"/modules/{module_id}/import/preview",
        files={"file": ("data.csv", content, "text/csv")},
        headers=_auth(token),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert set(body["column_names"]) == {"dim", "Sales", "COGS"}
    assert len(body["sample_rows"]) == 1
    assert "suggested_mapping" in body
    assert body["suggested_mapping"]["dim"] is None


# ── 23. POST /dimensions/{id}/import — bad column name → 422 ─────────────────

@pytest.mark.asyncio
async def test_dimension_import_bad_column(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    dim_id = await _create_dimension(client, token, model_id)

    content = _make_csv([{"name": "Widget A"}], ["name"])
    resp = await client.post(
        f"/dimensions/{dim_id}/import",
        files={"file": ("data.csv", content, "text/csv")},
        params={"name_column": "nonexistent_column"},
        headers=_auth(token),
    )
    assert resp.status_code == 422


# ── 24. POST /modules/{id}/import — no matching line-item columns → 422 ───────

@pytest.mark.asyncio
async def test_module_import_no_matching_columns(client: AsyncClient):
    token = await _register_login(client)
    ws_id = await _create_workspace(client, token)
    model_id = await _create_model(client, token, ws_id)
    module_id = await _create_module(client, token, model_id)
    await _create_line_item(client, token, module_id, "Sales")

    # Columns "FooBar" and "Baz" don't match any line item
    content = _make_csv([{"FooBar": "100", "Baz": "50"}], ["FooBar", "Baz"])
    resp = await client.post(
        f"/modules/{module_id}/import",
        files={"file": ("data.csv", content, "text/csv")},
        headers=_auth(token),
    )
    assert resp.status_code == 422


# ── 25. GET /modules/{id}/export — unknown module → 404 ──────────────────────

@pytest.mark.asyncio
async def test_module_export_not_found(client: AsyncClient):
    token = await _register_login(client)
    fake_id = uuid.uuid4()
    resp = await client.get(
        f"/modules/{fake_id}/export",
        params={"format": "csv"},
        headers=_auth(token),
    )
    assert resp.status_code == 404
