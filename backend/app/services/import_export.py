"""
Import/export service: CSV and Excel parsing, dimension import,
module cell import, and module export.
"""
import csv
import io
import uuid
from typing import Any, Dict, List, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cell import CellValue
from app.models.dimension import Dimension, DimensionItem, DimensionType
from app.models.module import LineItem, Module
from app.schemas.import_export import ImportResult
from app.services.cell import write_cell


# ── CSV parsing ────────────────────────────────────────────────────────────────

def parse_csv(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse CSV bytes into a list of row dicts (keyed by column header)."""
    text = file_content.decode("utf-8-sig")  # handle BOM
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for row in reader:
        rows.append(dict(row))
    return rows


# ── Excel parsing ──────────────────────────────────────────────────────────────

def parse_excel(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse Excel (.xlsx) bytes into a list of row dicts using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # First row is the header
            header_row = [str(cell) if cell is not None else f"col_{j}" for j, cell in enumerate(row)]
        else:
            if header_row is None:
                break
            row_dict: Dict[str, Any] = {}
            for col_name, cell_value in zip(header_row, row):
                row_dict[col_name] = cell_value
            # Skip entirely empty rows
            if any(v is not None and str(v).strip() != "" for v in row_dict.values()):
                rows.append(row_dict)

    return rows


# ── Dimension import ───────────────────────────────────────────────────────────

async def import_to_dimension(
    db: AsyncSession,
    dimension_id: uuid.UUID,
    rows: List[Dict[str, Any]],
    name_column: str,
    parent_column: Optional[str] = None,
) -> ImportResult:
    """
    Import dimension items from parsed rows.

    - name_column: column in each row that holds the item name/code.
    - parent_column: optional column that holds the parent item name/code.
    Returns an ImportResult with counts and errors.
    """
    rows_imported = 0
    rows_skipped = 0
    errors: List[str] = []

    dimension_result = await db.execute(
        select(Dimension).where(Dimension.id == dimension_id)
    )
    dimension = dimension_result.scalar_one_or_none()
    if dimension is None:
        return ImportResult(
            rows_imported=0,
            rows_skipped=len(rows),
            errors=["Dimension not found"],
        )

    is_numbered_dimension = dimension.dimension_type == DimensionType.numbered

    # Load existing items so we can resolve parent references
    result = await db.execute(
        select(DimensionItem).where(DimensionItem.dimension_id == dimension_id)
    )
    existing_items = list(result.scalars().all())
    # name -> item  (for parent lookup)
    name_to_item: Dict[str, DimensionItem] = {item.name: item for item in existing_items}

    # Determine the starting sort_order
    sort_base = max((item.sort_order for item in existing_items), default=-1) + 1
    next_numbered_code = 1
    if is_numbered_dimension:
        existing_numeric_codes: List[int] = []
        for item in existing_items:
            try:
                existing_numeric_codes.append(int(item.code))
            except (TypeError, ValueError):
                continue
        next_numbered_code = (max(existing_numeric_codes) + 1) if existing_numeric_codes else 1

    for idx, row in enumerate(rows):
        raw_name = row.get(name_column)
        if raw_name is None or str(raw_name).strip() == "":
            rows_skipped += 1
            errors.append(f"Row {idx + 1}: missing value for column '{name_column}', skipped")
            continue

        name = str(raw_name).strip()

        if is_numbered_dimension:
            projected_count = len(existing_items) + rows_imported
            if (
                dimension.max_items is not None
                and projected_count >= dimension.max_items
            ):
                rows_skipped += 1
                errors.append(
                    f"Row {idx + 1}: numbered list max_items ({dimension.max_items}) reached, skipped"
                )
                continue
            code = str(next_numbered_code)
            next_numbered_code += 1
        else:
            code = name  # use name as code (unique within dimension)

        parent_id: Optional[uuid.UUID] = None
        if parent_column:
            raw_parent = row.get(parent_column)
            if raw_parent is not None and str(raw_parent).strip() != "":
                parent_name = str(raw_parent).strip()
                parent_item = name_to_item.get(parent_name)
                if parent_item is None:
                    errors.append(
                        f"Row {idx + 1}: parent '{parent_name}' not found in dimension, item imported without parent"
                    )
                else:
                    parent_id = parent_item.id

        # For standard lists, skip duplicate names. Numbered lists can reuse
        # the same display name across multiple items.
        if not is_numbered_dimension and name in name_to_item:
            rows_skipped += 1
            errors.append(f"Row {idx + 1}: item '{name}' already exists, skipped")
            continue

        new_item = DimensionItem(
            name=name,
            code=code,
            dimension_id=dimension_id,
            parent_id=parent_id,
            sort_order=sort_base + rows_imported,
        )
        db.add(new_item)
        # Register so subsequent rows can reference this as a parent
        name_to_item[name] = new_item
        rows_imported += 1

    await db.commit()
    return ImportResult(rows_imported=rows_imported, rows_skipped=rows_skipped, errors=errors)


# ── Module cell import ─────────────────────────────────────────────────────────

async def import_to_module(
    db: AsyncSession,
    module_id: uuid.UUID,
    rows: List[Dict[str, Any]],
    line_item_mapping: Dict[str, uuid.UUID],
    dimension_mapping: Dict[str, uuid.UUID],
) -> ImportResult:
    """
    Import cell values from parsed rows into a module.

    - line_item_mapping: maps CSV column name -> LineItem UUID.
    - dimension_mapping: maps CSV column name -> DimensionItem UUID to use
      as the dimension key component for each row.
    Returns an ImportResult.
    """
    rows_imported = 0
    rows_skipped = 0
    errors: List[str] = []

    # Build dimension member list (same for all rows in this import)
    dimension_members = list(dimension_mapping.values())

    for idx, row in enumerate(rows):
        # For each mapped line item column, write a cell
        row_had_data = False
        for col_name, line_item_id in line_item_mapping.items():
            raw_value = row.get(col_name)
            if raw_value is None or str(raw_value).strip() == "":
                continue

            # Attempt to coerce to number
            value: Any = raw_value
            try:
                float_val = float(str(raw_value).replace(",", ""))
                value = float_val
            except (ValueError, TypeError):
                value = str(raw_value).strip()

            try:
                await write_cell(
                    db=db,
                    line_item_id=line_item_id,
                    dimension_members=dimension_members,
                    version_id=None,
                    value=value,
                )
                row_had_data = True
            except Exception as exc:
                errors.append(f"Row {idx + 1}, column '{col_name}': {exc}")

        if row_had_data:
            rows_imported += 1
        else:
            rows_skipped += 1

    return ImportResult(rows_imported=rows_imported, rows_skipped=rows_skipped, errors=errors)


# ── Module export ──────────────────────────────────────────────────────────────

async def _load_module_data(
    db: AsyncSession,
    module_id: uuid.UUID,
) -> tuple:
    """
    Load module, its line items, and all cell values.

    Returns (module, line_items, cells_by_line_item).
    """
    module_result = await db.execute(
        select(Module).where(Module.id == module_id)
    )
    module = module_result.scalar_one_or_none()
    if module is None:
        return None, [], {}

    li_result = await db.execute(
        select(LineItem).where(LineItem.module_id == module_id)
    )
    line_items: List[LineItem] = list(li_result.scalars().all())

    cells_by_line_item: Dict[uuid.UUID, List[CellValue]] = {}
    for li in line_items:
        cv_result = await db.execute(
            select(CellValue).where(CellValue.line_item_id == li.id)
        )
        cells_by_line_item[li.id] = list(cv_result.scalars().all())

    return module, line_items, cells_by_line_item


def _build_export_rows(
    line_items: List[Any],
    cells_by_line_item: Dict[uuid.UUID, List[Any]],
) -> List[Dict[str, Any]]:
    """
    Build a flat list of dicts suitable for CSV/Excel export.

    Each row represents one cell. Columns: dimension_key, line_item_name, value.
    """
    export_rows: List[Dict[str, Any]] = []

    for li in line_items:
        cells = cells_by_line_item.get(li.id, [])
        for cell in cells:
            if cell.value_boolean is not None:
                value: Any = cell.value_boolean
            elif cell.value_number is not None:
                value = cell.value_number
            elif cell.value_text is not None:
                value = cell.value_text
            else:
                value = None

            export_rows.append({
                "line_item": li.name,
                "dimension_key": cell.dimension_key,
                "value": value,
            })

    return export_rows


async def export_module_to_csv(db: AsyncSession, module_id: uuid.UUID) -> bytes:
    """Export module data (all line items and cell values) to CSV bytes."""
    module, line_items, cells_by_line_item = await _load_module_data(db, module_id)

    output = io.StringIO()
    fieldnames = ["line_item", "dimension_key", "value"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    rows = _build_export_rows(line_items, cells_by_line_item)
    for row in rows:
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


async def export_module_to_excel(db: AsyncSession, module_id: uuid.UUID) -> bytes:
    """Export module data (all line items and cell values) to Excel (.xlsx) bytes."""
    import openpyxl
    from openpyxl.styles import Font

    module, line_items, cells_by_line_item = await _load_module_data(db, module_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    module_name = module.name if module else "Module"
    ws.title = module_name[:31]  # Excel sheet name limit

    headers = ["line_item", "dimension_key", "value"]
    ws.append(headers)
    # Bold the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = _build_export_rows(line_items, cells_by_line_item)
    for row in rows:
        ws.append([row["line_item"], row["dimension_key"], row["value"]])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ── Import preview ─────────────────────────────────────────────────────────────

def build_import_preview(
    rows: List[Dict[str, Any]],
    max_sample: int = 5,
) -> Dict[str, Any]:
    """
    Build a preview dict from parsed rows.

    Returns column_names, sample_rows, and a suggested_mapping
    (all columns mapped to None — the caller fills in the mapping).
    """
    if not rows:
        return {
            "column_names": [],
            "sample_rows": [],
            "suggested_mapping": {},
        }

    column_names = list(rows[0].keys())
    sample_rows = rows[:max_sample]
    suggested_mapping: Dict[str, Optional[str]] = {col: None for col in column_names}

    return {
        "column_names": column_names,
        "sample_rows": sample_rows,
        "suggested_mapping": suggested_mapping,
    }
